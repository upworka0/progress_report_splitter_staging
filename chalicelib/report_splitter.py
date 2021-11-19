import base64
import io
import logging
import datetime
import openpyxl
import os
import pandas as pd
import zipfile
import requests
from copy import deepcopy
from openpyxl.styles import PatternFill
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (Mail, Attachment, FileContent, FileName, FileType, Disposition)
import bugsnag


'''
    Report Splitter

    Convert CSV files into Excels and send emails to recipients with converted excel and pdf files
'''
bucket_name2 = os.environ.get('BUCKET_NAME2', 'getinclusive-ppj-reports')
pdf_report_endpoint = os.environ.get('PDF_REPORT_LINK', 'https://sf72v4f1ib.execute-api.us-east-1.amazonaws.com/api/pdf-reports')


class ReportSplitter:
    group_by = None
    sender = ''
    recipient_list = []
    email_subject = ''
    email_body_data = {}
    email_body = ''

    report_gen_cell_dict = None
    report_name = ''
    filter_message = ''
    course_id = ''
    course_name = ''

    attach_files = []
    generated_timestamp = None
    output_folder_key = None

    total_size = 0
    curr_dt_tm = None

    from_zip = False    # Flag to determine csv is from zip or not.

    zip_folder_path = None

    def __init__(self, app, s3, bucket_name, input_template_key, sendgrid_key, bugsnag_key):
        self.app = app
        self.s3 = s3
        self.bucket_name = bucket_name
        self.input_template_key = input_template_key
        self.sendgrid_key = sendgrid_key
        self.bugsnag_key = bugsnag_key

    def bugsnag_handler(self, e):
        client = bugsnag.Client(api_key=self.bugsnag_key)
        client.notify(e)

    def logger(self, msg, ty='info'):
        if ty == 'error':
            self.app.log.error(msg)
        else:
            self.app.log.info(msg)

    def current_dt_tm(self):
        return datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

    # this method takes a df and return the result file objs using the template
    def data_cleaner(self, df, index, output_folder_key, report_gen_cell_dict, input_df_col, report_name):
        self.logger("Inside cleaner 1")
        input_template_obj = self.s3.get_object(Bucket=self.bucket_name, Key=self.input_template_key)
        input_template_data = input_template_obj['Body'].read()
        template_book = openpyxl.load_workbook(io.BytesIO(input_template_data))
        self.logger(input_df_col)
        for x in template_book.sheetnames:
            t_sheet = template_book[x]
            for y in report_gen_cell_dict.keys():
                for z in report_gen_cell_dict[y]:
                    if type(z[1]) is not str:
                        z[1] = str(z[1])
                    t_sheet[y] = str(t_sheet[y].value).replace(z[0], z[1])

        self.logger("Inside cleaner 2")
        pending_df = deepcopy(df.loc[(df['completion_percent'] < 100) & (df['Active?'] == True)])
        pending_df.fillna('', inplace=True)

        pending_sheet = template_book["Pending"]
        self.logger("Inside cleaner 3")
        for i in range(1, len(input_df_col) + 1):
            pending_sheet.cell(row=6,
                               column=i,
                               value=input_df_col[i - 1])

            _cell = pending_sheet.cell(row=6, column=i)
            _cell.fill = PatternFill("solid", fgColor="D3D3D3")

        pending_df_starts_row = 7
        pending_df_starts_column = 1
        self.logger("Inside cleaner 4")
        for r_idx in range(0, pending_df.shape[0]):
            for c_idx in range(0, pending_df.shape[1]):
                value = pending_df.iat[r_idx, c_idx]
                # remove all brackets and quote symbols except reference_data field
                if type(value) is str and c_idx != 11:
                    value = value.replace("[", '').replace("]", '').replace('"', '')
                pending_sheet.cell(row=r_idx + pending_df_starts_row,
                                   column=c_idx + pending_df_starts_column,
                                   value=value)

        completed_df = deepcopy(df.loc[(df['completion_percent'] == 100) & (df['Active?'] == True)])
        completed_df.fillna('', inplace=True)

        self.logger("Inside cleaner 5")
        completed_sheet = template_book["Completed"]

        for i in range(1, len(input_df_col) + 1):
            completed_sheet.cell(row=6,
                                 column=i,
                                 value=input_df_col[i - 1])

            _cell = completed_sheet.cell(row=6, column=i)
            _cell.fill = PatternFill("solid", fgColor="D3D3D3")
        self.logger("Inside cleaner 6")
        completed_df_starts_row = 7
        completed_df_starts_column = 1
        for r_idx in range(0, completed_df.shape[0]):
            for c_idx in range(0, completed_df.shape[1]):
                value = completed_df.iat[r_idx, c_idx]
                # remove all brackets and quote symbols except reference_data field
                if type(value) is str and c_idx != 11:
                    value = value.replace("[", '').replace("]", '').replace('"', '')
                completed_sheet.cell(row=r_idx + completed_df_starts_row,
                                     column=c_idx + completed_df_starts_column,
                                     value=value)

        self.logger("Inside cleaner 7")
        # deactivated_df = deepcopy(df.loc[(df['completion_percent'] < 100) & (df['Active?'] == False)])
        deactivated_df = deepcopy(df.loc[df['Active?'] == False])
        deactivated_df.fillna('', inplace=True)
        # deactivated_df.loc[-1] = list(deactivated_df.columns)
        # deactivated_df.sort_index(inplace=True)
        # deactivated_df.reset_index(inplace=True, drop=True)

        deactivated_sheet = template_book["Deactivated"]

        for i in range(1, len(input_df_col) + 1):
            deactivated_sheet.cell(row=6,
                                   column=i,
                                   value=input_df_col[i - 1])

            _cell = deactivated_sheet.cell(row=6, column=i)
            _cell.fill = PatternFill("solid", fgColor="D3D3D3")

        self.logger("Inside cleaner 8")
        deactivated_df_starts_row = 7
        deactivated_df_starts_column = 1
        for r_idx in range(0, deactivated_df.shape[0]):
            for c_idx in range(0, deactivated_df.shape[1]):
                value = deactivated_df.iat[r_idx, c_idx]
                # remove all brackets and quote symbols except reference_data field
                if type(value) is str and c_idx != 11:
                    value = value.replace("[", '').replace("]", '').replace('"', '')
                deactivated_sheet.cell(row=r_idx + deactivated_df_starts_row,
                                       column=c_idx + deactivated_df_starts_column,
                                       value=value)

        output_file_memory_obj = io.BytesIO()
        template_book.save(output_file_memory_obj)
        output_file_memory_obj.seek(0)
        self.logger("Inside cleaner 9")
        self.logger(index)
        if isinstance(index, (str, int, float)):
            index = [index]

        if self.from_zip:
            output_file_xlsx = "{}--{}--{}.xlsx".format(report_name, self.course_name, "--".join(map(str, index)))
        else:
            output_file_xlsx = "{}--{}.xlsx".format(report_name, "--".join(map(str, index)))

        self.s3.put_object(Bucket=self.bucket_name,
                           Key=output_folder_key + output_file_xlsx,
                           Body=output_file_memory_obj)

        self.logger("Inside cleaner 10")

        self.total_size += output_file_memory_obj.__sizeof__()
        return output_file_xlsx, output_file_memory_obj

    # check if the total size of xlsx files is more than 7MB, and compress them as zip file
    def zip_attachments(self, bucket, output_folder):
        self.logger('Zipping files ...')
        zip_data = io.BytesIO()
        output_file_name = "{}_{}.zip".format(self.report_name, self.curr_dt_tm)
        new_attach_files = []

        with zipfile.ZipFile(zip_data, mode='a') as zf:
            for attach_file in self.attach_files:
                if attach_file['type'] == 'application/vnd.ms-excel':
                    self.app.log.info(attach_file['key'])
                    attach_file['data'].seek(0)
                    zf.writestr(attach_file['key'], attach_file['data'].read())
                else:
                    new_attach_files.append(attach_file)

        self.attach_files = new_attach_files

        # put the zip file to s3 bucket
        zip_data.seek(0)
        self.s3.put_object(Bucket=bucket,
                           Key=output_folder + output_file_name,
                           Body=zip_data)

        self.logger('Zipping end. {}{}'.format(output_folder, output_file_name))
        return output_folder + output_file_name

    def send_email_with_attachment(self):
        self.logger('Email sending ...')
        self.logger('Recipient_list ...')
        self.logger(self.recipient_list)
        try:
            message = Mail(
                from_email=self.sender,
                to_emails=self.recipient_list
            )
            attachment = "Y"
            file_link = ""

            if self.total_size > 10000000:
                file_key = self.zip_attachments(self.bucket_name, self.output_folder_key)
                file_link = self.s3.generate_presigned_url(
                    'get_object',
                    Params={'Bucket': self.bucket_name, 'Key': file_key},
                    ExpiresIn=604800
                )
                attachment = ""

            for attach_file in self.attach_files:
                print(attach_file['key'])
                data = attach_file['data']
                data.seek(0)
                encoded_file = base64.b64encode(data.read()).decode()

                attached_file = Attachment(
                    FileContent(encoded_file),
                    FileName(attach_file['key']),
                    FileType(attach_file['type']),
                    Disposition('attachment')
                )
                message.add_attachment(attached_file)

            message.dynamic_template_data = {
                'progress_report_date': self.email_body_data['generated_timestamp'],
                'attachment': attachment,
                'link': file_link,
                'subject': self.email_subject,
            }
            message.template_id = 'd-0bde5f7432074190a571e7d7233b1416'

            sg = SendGridAPIClient(self.sendgrid_key)
            response = sg.send(message)

            self.logger(response.status_code)
            self.logger(response.body)

            """ Compress all files as zip format and upload to getinclusive-ppj-reports bucket for monitoring bugs """
            self.logger('Zipping and uploading on {}'.format(bucket_name2))
            self.zip_attachments(bucket_name2, self.zip_folder_path)
            self.logger('Success --- END')
            return None

        except Exception as e:
            self.logger("Error from sending email:" + str(e), ty="error")
            self.bugsnag_handler(e)
            return {
                'code': 400,
                'status': 'error',
                'error_msg': "Error from sending email:" + str(e)
            }

    # Process a csv file with its byte data
    def process_csv(self, input_csv_data):
        self.report_gen_cell_dict = {
            'A2': [['{{course_alias}}', self.course_name]],
            'A3': [
                ['{{timestamp}}', self.generated_timestamp],
                ['{{course_id}}', self.course_id]
            ],
            'A4': [['{{filter_abc}}', self.filter_message]]
        }

        input_df = pd.read_csv(io.BytesIO(input_csv_data), low_memory=False)
        input_df['Active?'].replace('(?i)true', True, inplace=True, regex=True)
        input_df['Active?'].replace('(?i)false', False, inplace=True, regex=True)
        input_df['Active?'].fillna(True, inplace=True)
        input_df['completion_percent'].fillna(0, inplace=True)
        input_df['completion_percent'] = input_df['completion_percent'].astype(float)

        input_df_col = list(input_df.columns)

        if self.group_by:
            input_df_grouped = input_df.groupby(self.group_by)
            for count, (index, df) in enumerate(input_df_grouped):
                output_file_xlsx, output_file_memory_obj = self.data_cleaner(
                    df, index, self.output_folder_key, self.report_gen_cell_dict, input_df_col,
                    self.report_name
                )
                self.attach_files.append({
                    'key': output_file_xlsx,
                    'type': 'application/vnd.ms-excel',
                    'data': output_file_memory_obj
                })
        else:
            '''
            data cleaner will return xlsx file name and file obj. Unless group by, there will be only one xlsx file
            no zip is necessary
            '''
            output_file_xlsx, output_file_memory_obj = self.data_cleaner(
                input_df, 'FULL_FILE', self.output_folder_key, self.report_gen_cell_dict, input_df_col,
                self.report_name
            )
            self.attach_files.append({
                'key': output_file_xlsx,
                'type': 'application/vnd.ms-excel',
                'data': output_file_memory_obj
            })

    def process(self, input_dict, event_key=''):
        self.curr_dt_tm = self.current_dt_tm()
        self.output_folder_key = 'progress_reports/{}/output_folder_{}/'.format(input_dict['org_id'], self.curr_dt_tm)
        self.filter_message = input_dict.get('filter_message', '')
        self.report_name = str(input_dict.get('report_name', ''))
        self.generated_timestamp = input_dict.get('generated_timestamp', '')
        self.email_body = input_dict.get('email_body', '')
        self.email_subject = input_dict.get('email_subject', 'Assignment Progress Report')
        self.group_by = input_dict.get('group_by', '')
        self.recipient_list = input_dict.get('email_addresses', [])
        self.sender = input_dict.get('email_sender', '')

        self.zip_folder_path = 'progress_reports/{}/{} {}/'.format(input_dict['org_id'], " ".join(self.recipient_list), self.curr_dt_tm)
        self.s3.put_object(Bucket=bucket_name2, Key=self.zip_folder_path)

        self.email_body_data = {
            'generated_timestamp': self.generated_timestamp,
            'email_body': self.email_body,
            'email_subject': self.email_subject,
        }

        self.attach_files = []

        try:
            # Process with CSV file
            if 'CSV_filepath' in input_dict and input_dict['CSV_filepath']:
                input_csv_key = 'progress_reports/{}/{}'.format(input_dict['org_id'], input_dict['CSV_filepath'])
                self.logger('Input CSV file is %s' % input_csv_key)
                input_csv_obj = self.s3.get_object(Bucket=self.bucket_name, Key=input_csv_key)
                input_csv_data = input_csv_obj['Body'].read()

                # Retrieve CourseID and CourseName from InputDict
                self.course_name = input_dict.get('course_name', '')
                self.course_id = input_dict.get('course_id', '')
                self.from_zip = False
                self.process_csv(input_csv_data)
            # Process with Zip file
            elif 'ZIP_filepath' in input_dict and input_dict['ZIP_filepath']:
                input_zip_key = 'progress_reports/{}/{}'.format(input_dict['org_id'], input_dict['ZIP_filepath'])
                self.logger('Input Zip file is %s' % input_zip_key)
                course_ids = []     # available course_id list
                input_zip_obj = self.s3.get_object(Bucket=self.bucket_name, Key=input_zip_key)
                input_zip_buf = io.BytesIO(input_zip_obj['Body'].read())
                z = zipfile.ZipFile(input_zip_buf)
                self.from_zip = True
                for filename in z.namelist():
                    if filename.endswith('.csv'):
                        input_csv_data = z.read(filename)

                        # Retrieve CourseID and CourseName from Courses list
                        self.course_name = input_dict['courses'][filename].get('course_name')
                        self.course_id = input_dict['courses'][filename].get('course_id', '')

                        self.process_csv(input_csv_data)
                        course_ids.append(str(self.course_id))
                    else:
                        # Append files if it's not zip or csv type
                        _, file_type = os.path.splitext(filename)
                        if 'pdf' not in file_type:
                            self.attach_files.append({
                                'key': f'OLD - {filename}',
                                'type': file_type,
                                'data': io.BytesIO(z.read(filename))
                            })

                # append report pdf to attachment.
                pdf_url = "{}?org_id={}&course_ids={}".format(pdf_report_endpoint, input_dict['org_id'], ",".join(course_ids))
                res = requests.get(pdf_url)
                pdf_data = io.BytesIO(res.content)
                filename = res.headers['content-disposition'].split("filename=")[1]
                self.attach_files.append({
                    'key': filename,
                    'type': 'pdf',
                    'data': pdf_data
                })
            else:
                # TODO: if there is no CSV or Zip file path in input_dict, need more discussion
                pass
        except Exception as e:
            self.logger("Error: " + str(e), ty='error')
            self.bugsnag_handler(e)
            return {
                'code': 400,
                'status': 'error',
                'error_msg': str(e)
            }

        res = self.send_email_with_attachment()  # send email with converted files
        if res:
            return res

        if event_key:
            try:
                copy_source = {
                    'Bucket': self.bucket_name,
                    'Key': event_key
                }
                self.s3.copy(copy_source, self.bucket_name,
                             'progress_reports/{}/{}'.format(input_dict['org_id'], event_key.split('/')[-1]))
                self.s3.delete_object(Bucket=self.bucket_name, Key=event_key)
            except Exception as e:
                self.bugsnag_handler(e)
                self.logger("Error: " + str(e), ty='error')

        return {
            'code': 200,
            'status': 'ok',
            'output_folder': self.output_folder_key
        }

