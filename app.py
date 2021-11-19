import base64
import boto3
import datetime
import io
import logging
import openpyxl
import os
import pandas as pd
import json
import zipfile
import bugsnag
import requests
from chalice import Chalice
from copy import deepcopy
from openpyxl.styles import PatternFill
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (Mail, Attachment, FileContent, FileName, FileType, Disposition)
from chalicelib.report_splitter import ReportSplitter

app = Chalice(app_name='progress-report-generator')

# Enable DEBUG logs.
app.log.setLevel(logging.DEBUG)

s3 = boto3.client('s3')
sqs = boto3.resource('sqs')

bucket_name = os.environ.get('BUCKET_NAME', 'getinclusive-ppj-private')
input_template_key = os.environ.get('INPUT_TEMPLATE_KEY', 'progress_reports/Template.xlsx')
sendgrid_key = os.environ.get('SENDGRID_KEY')
queue_name = os.environ.get('QUEUE_NAME')
bugsnag_key = os.environ.get('BUGSNAG_KEY')

# configure bugsnag
bugsnag.configure(api_key=bugsnag_key)


def current_dt_tm():
    return datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")


def send_email_attachment(
        attachment_path, email_body_data, sender='', recipient_list=[],
        zip_report_name='data.zip', file_key='', email_subject='', zip_flag=True
):
    try:
        message = Mail(
            from_email=sender,
            to_emails=recipient_list
        )

        attachment_path.seek(0)
        encoded_file = base64.b64encode(attachment_path.read()).decode()

        attachment_file_size = attachment_path.__sizeof__()
        # print(attachment_file_size)
        # breakpoint()
        # less then 10 MB

        if zip_flag:
            attachment_file_type = 'application/zip'
        else:
            attachment_file_type = 'application/vnd.ms-excel'

        if attachment_file_size < 10000000:
            # if False:
            attached_file = Attachment(
                FileContent(encoded_file),
                FileName(zip_report_name),
                FileType(attachment_file_type),
                Disposition('attachment')
            )
            message.attachment = attached_file
            attachment = "Y"
            file_link = ""
        else:
            attachment = ""
            file_link = s3.generate_presigned_url('get_object',
                                                  Params={'Bucket': bucket_name,
                                                          'Key': file_key},
                                                  ExpiresIn=604800)

        # print(file_link)
        # breakpoint()
        # pass custom values for our HTML placeholders
        message.dynamic_template_data = {
            'progress_report_date': email_body_data['generated_timestamp'],
            'attachment': attachment,
            'link': file_link,
            'subject': email_subject
        }
        message.template_id = 'd-0bde5f7432074190a571e7d7233b1416'

        sg = SendGridAPIClient(os.environ.get('SENDGRID_KEY'))
        response = sg.send(message)

        app.log.info(response.status_code)
        # app.log.info(response.body)
        # app.log.info(response.headers)

    except Exception as e:
        app.log.info("Error from sending email:" + str(e))


# this method takes a df and return the result file objs using the template
def data_cleaner(df, index, output_folder_key, report_gen_cell_dict, input_df_col, report_name):
    app.log.info("Inside cleaner 1")
    input_template_obj = s3.get_object(Bucket=bucket_name, Key=input_template_key)
    input_template_data = input_template_obj['Body'].read()
    template_book = openpyxl.load_workbook(io.BytesIO(input_template_data))
    app.log.info(input_df_col)
    for x in template_book.sheetnames:
        t_sheet = template_book[x]
        for y in report_gen_cell_dict.keys():
            for z in report_gen_cell_dict[y]:
                t_sheet[y] = str(t_sheet[y].value).replace(z[0], z[1])

        # need to capitalize the
        # print(t_sheet.title)
        # t_sheet.title = str(t_sheet.title).capitalize()

    # print(template_book.sheetnames)
    # breakpoint()
    # this template has 3 sheet and starts from A4
    # sheet_name = ['pending', 'completed', 'deactivated']
    app.log.info("Inside cleaner 2")
    pending_df = deepcopy(df.loc[(df['completion_percent'] < 100) & (df['Active?'] == True)])
    pending_df.fillna('', inplace=True)
    # pending_df.loc[-1] = list(pending_df.columns)
    # pending_df.sort_index(inplace=True)
    # pending_df.reset_index(inplace=True, drop=True)

    pending_sheet = template_book["Pending"]
    app.log.info("Inside cleaner 3")
    for i in range(1, len(input_df_col) + 1):
        pending_sheet.cell(row=6,
                           column=i,
                           value=input_df_col[i - 1])

        _cell = pending_sheet.cell(row=6, column=i)
        _cell.fill = PatternFill("solid", fgColor="D3D3D3")

    pending_df_starts_row = 7
    pending_df_starts_column = 1
    app.log.info("Inside cleaner 4")
    for r_idx in range(0, pending_df.shape[0]):
        for c_idx in range(0, pending_df.shape[1]):
            pending_sheet.cell(row=r_idx + pending_df_starts_row,
                               column=c_idx + pending_df_starts_column,
                               value=pending_df.iat[r_idx, c_idx])

    completed_df = deepcopy(df.loc[(df['completion_percent'] == 100) & (df['Active?'] == True)])
    completed_df.fillna('', inplace=True)
    # completed_df.loc[-1] = list(completed_df.columns)
    # completed_df.sort_index(inplace=True)
    # completed_df.reset_index(inplace=True, drop=True)
    app.log.info("Inside cleaner 5")
    completed_sheet = template_book["Completed"]

    for i in range(1, len(input_df_col) + 1):
        completed_sheet.cell(row=6,
                             column=i,
                             value=input_df_col[i - 1])

        _cell = completed_sheet.cell(row=6, column=i)
        _cell.fill = PatternFill("solid", fgColor="D3D3D3")
    app.log.info("Inside cleaner 6")
    completed_df_starts_row = 7
    completed_df_starts_column = 1
    for r_idx in range(0, completed_df.shape[0]):
        for c_idx in range(0, completed_df.shape[1]):
            completed_sheet.cell(row=r_idx + completed_df_starts_row,
                                 column=c_idx + completed_df_starts_column,
                                 value=completed_df.iat[r_idx, c_idx])

    app.log.info("Inside cleaner 7")
    # deactivated_df = deepcopy(df.loc[(df['completion_percent'] < 100) & (df['Active?'] == False)])
    deactivated_df = deepcopy(df.loc[df['Active?'] == False])
    deactivated_df.fillna('', inplace=True)
    # deactivated_df.loc[-1] = list(deactivated_df.columns)
    # deactivated_df.sort_index(inplace=True)
    # deactivated_df.reset_index(inplace=True, drop=True)

    deactivated_sheet = template_book["Deactivated"]

    for i in range(1, len(input_df_col) + 1):
        deactivated_sheet.cell(row=6,
                               column=i,
                               value=input_df_col[i - 1])

        _cell = deactivated_sheet.cell(row=6, column=i)
        _cell.fill = PatternFill("solid", fgColor="D3D3D3")

    app.log.info("Inside cleaner 8")
    deactivated_df_starts_row = 7
    deactivated_df_starts_column = 1
    for r_idx in range(0, deactivated_df.shape[0]):
        for c_idx in range(0, deactivated_df.shape[1]):
            deactivated_sheet.cell(row=r_idx + deactivated_df_starts_row,
                                   column=c_idx + deactivated_df_starts_column,
                                   value=deactivated_df.iat[r_idx, c_idx])

    output_file_memory_obj = io.BytesIO()
    template_book.save(output_file_memory_obj)
    output_file_memory_obj.seek(0)
    app.log.info("Inside cleaner 9")
    app.log.info(index)
    if isinstance(index, (str, int, float)):
        index = [index]
    output_file_xlsx = str(report_name) + '--' + "--".join(map(str, index)) + '.xlsx'
    s3.put_object(Bucket=bucket_name,
                  Key=output_folder_key + output_file_xlsx,
                  Body=output_file_memory_obj)

    app.log.info("Inside cleaner 10")
    return output_file_xlsx, output_file_memory_obj


def excel_splitter_init(input_dict, event_key=''):
    curr_dt_tm = current_dt_tm()
    try:
        input_csv_key = 'progress_reports/' + str(input_dict['org_id']) + '/' + input_dict['CSV_filepath']
        input_csv_obj = s3.get_object(Bucket=bucket_name, Key=input_csv_key)
        input_csv_data = input_csv_obj['Body'].read()

        # app.log.debug(input_csv_data)
        # app.log.debug(type(input_csv_data))

        input_df = pd.read_csv(io.BytesIO(input_csv_data))
        input_df['Active?'].replace('(?i)true', True, inplace=True, regex=True)
        input_df['Active?'].replace('(?i)false', False, inplace=True, regex=True)
        input_df['Active?'].fillna(True, inplace=True)
        input_df['completion_percent'].fillna(0, inplace=True)
        input_df['completion_percent'] = input_df['completion_percent'].astype(float)
        # app.log.debug(input_df.dtypes)
        # breakpoint()

        output_folder_key = 'progress_reports/' + str(input_dict['org_id']) + '/' + 'output_folder_' + curr_dt_tm + '/'

        course_name = input_dict.get('course_name', '')
        course_id = input_dict.get('course_id', '')
        filter_message = input_dict.get('filter_message', '')
        report_name = str(input_dict.get('report_name', ''))

        generated_timestamp = input_dict.get('generated_timestamp', '')
        email_body = input_dict.get('email_body', '')
        email_subject = input_dict.get('email_subject', 'Assignment Progress Report')

        email_body_data = {
            'generated_timestamp': generated_timestamp,
            'email_body': email_body,
            'email_subject': email_subject,
        }

        report_gen_cell_dict = {
            'A2': [['{{course_alias}}', course_name]],
            'A3': [
                ['{{timestamp}}', generated_timestamp],
                ['{{course_id}}', course_id]
            ],
            'A4': [['{{filter_abc}}', filter_message]]
        }

        input_df_col = list(input_df.columns)
        group_by = input_dict.get('group_by', '')
        output_file_path = io.BytesIO()
        zip_flag = True

        if group_by:
            input_df_grouped = input_df.groupby(group_by)
            output_file_name = report_name + '_' + curr_dt_tm + '.zip'
            for count, (index, df) in enumerate(input_df_grouped):
                output_file_xlsx, output_file_memory_obj = data_cleaner(
                    df, index, output_folder_key, report_gen_cell_dict, input_df_col,
                    report_name
                )

                with zipfile.ZipFile(output_file_path, mode='a') as zf:
                    output_file_memory_obj.seek(0)
                    zf.writestr(output_file_xlsx, output_file_memory_obj.read())
        else:
            '''
            data cleaner will return xlsx file name and file obj. Unless group by, there will be only one xlsx file
            no zip is necessary
            '''
            output_file_xlsx, output_file_memory_obj = data_cleaner(
                input_df, 'FULL_FILE', output_folder_key, report_gen_cell_dict, input_df_col,
                report_name
            )

            output_file_name = output_file_xlsx
            output_file_path = output_file_memory_obj
            zip_flag = False

        if zip_flag:
            output_file_path.seek(0)
            s3.put_object(Bucket=bucket_name,
                          Key=output_folder_key + output_file_name,
                          Body=output_file_path)

        send_email_attachment(output_file_path, sender=input_dict['email_sender'],
                              recipient_list=input_dict['email_addresses'],
                              zip_report_name=output_file_name,
                              email_body_data=email_body_data,
                              file_key=output_folder_key + output_file_name,
                              email_subject=email_subject,
                              zip_flag=zip_flag)

        # app.log.debug('progress_reports/' + str(input_dict['org_id']) + '/' + event_key.split('/')[-1])
        # app.log.debug(event_key)
    except Exception as e:
        app.log.error("Error: " + str(e))
        return {
            'code': 400,
            'status': 'error',
            'error_msg': str(e)
        }

    # the json file must be moved to org_folder from root
    # in api call, there will be no json file
    if event_key:
        try:
            copy_source = {
                'Bucket': bucket_name,
                'Key': event_key
            }
            s3.copy(copy_source, bucket_name,
                    'progress_reports/' + str(input_dict['org_id']) + '/' + event_key.split('/')[-1])
            s3.delete_object(Bucket=bucket_name, Key=event_key)
        except Exception as e:
            app.log.error("Error: " + str(e))

    return {
        'code': 200,
        'status': 'ok',
        'output_file': output_folder_key + report_name + '_' + curr_dt_tm + '.zip'
    }


'''
# Commentted because of error Configurations overlap of same bucket.
@app.on_s3_event(bucket=bucket_name)
def handler(event):
    # fire only when json file is created and placed in this sub dir
    # if not ("progress_reports" in event.key and event.key.endswith('.json')):
    #     return
    if re.search(r'progress_reports/split_report_(.*)\.json', event.key, re.MULTILINE | re.IGNORECASE) is None:
        return

    app.log.info("Handler triggered.....")
    app.log.info(datetime.datetime.now())
    app.log.info(event.key)
    app.log.info('regex passed')

    obj = s3.get_object(Bucket=bucket_name, Key=event.key)
    data = obj['Body']
    input_dict = json.load(data)

    json_response = excel_splitter_init(input_dict, event_key=event.key)

    return json_response
'''


@app.route('/', methods=['POST'])
def api_handler():
    header_data = app.current_request.headers
    if header_data.get('x-api-key', '').upper() != '7D6E97ED-E667-4C1F-9075-020CCC5C97EF':
        return {
            "status": "Error",
            "message": "Not Authorized, custom auth missing"
        }
    app.log.info("Handler triggered from api.....")
    request = app.current_request.raw_body.decode()
    input_dict = json.loads(request)
    json_response = excel_splitter_init(input_dict)

    return json_response


# '''
# tester, local
@app.route('/test')
def index():
    app.log.info("Handler triggered.....")
    #  fire only when json file is created and placed in this sub dir
    # if not ("progress_reports" in event.key and event.key.endswith('.json')):
    #     return
    # if re.search(r'progress_reports/split_report_(.*)\.json', event.key, re.MULTILINE | re.IGNORECASE) == None:
    #     return

    print('Program started....')
    event_key = 'progress_reports/test_org/split_report_test_org_no_group.json'
    event_key = 'progress_reports/test_org/split_report_test_org.json'
    # if re.search(r'progress_reports/split_report_(.*)\.json', event.key, re.MULTILINE | re.IGNORECASE) == None:
    #     return
    # event_key = event.key
    # event_key = 'progress_reports/1/split_report_xxxxxxx_2.json'
    # event_key = 'progress_reports/1/split_report_4Apr2021.json'
    obj = s3.get_object(Bucket=bucket_name, Key=event_key)
    data = obj['Body']
    input_dict = json.load(data)

    json_response = excel_splitter_init(input_dict, event_key=event_key)

    return json_response


# auth code testing
'''
from chalice import CustomAuthorizer
authorizer = CustomAuthorizer(
    'lambda_auth_with_api_key', header='Authorization',
    authorizer_uri=('arn:aws:lambda:us-east-1:333410091315:function:lambda_auth_with_api_key'))
@app.route('/test_auth', methods=['GET'], authorizer=authorizer)
def test_api_auth():
    header_data = app.current_request.headers
    if header_data.get('x-api-key', '') != '7D6E97ED-E667-4C1F-9075-020CCC5C97EF':
        return {    "status": "Error",
                    "message": "Not Authorized, custom auth missing"
            }
    return {    "status": "Ok1",
                "message": "Authorized!"
        }
'''


########################################################################################################################


'''
    Test new function

    Example JSON
    {
        "report_name": "ali_test",
        "org_id": "test_org",
        "course_name": "test_course_name",
        "course_id": "test_course_id",
        "filter_message": "test_filter_message",
        "generated_timestamp": "2021-1-22",
        "email_addresses": ["upworka0@gmail.com"],
        "email_sender": "upworka0@gmail.com",
        "CSV_filtpath": "report1_4Apr_50k.csv",
        "group_by": ["ref_State", "ref_City"],
        "ZIP_filepath": "sample_03.zip",
        "email_subject": "6 Apr testing"
    }
'''


@app.route('/new', methods=['POST'])
def new_handler():
    header_data = app.current_request.headers
    if header_data.get('x-api-key', '').upper() != '7D6E97ED-E667-4C1F-9075-020CCC5C97EF':
        return {
            "status": "Error",
            "message": "Not Authorized, custom auth missing"
        }

    request_body = app.current_request.raw_body.decode()
    app.log.info("Handler triggered from api.....")
    app.log.info('Request ...')
    app.log.info('Request body: %s' % request_body)

    try:
        queue = sqs.get_queue_by_name(QueueName=queue_name)
        response = queue.send_message(MessageBody=request_body)
        app.log.info(response)
    except Exception as e:
        app.log.error("Error: " + str(e))
        bugsnag.notify(e)
        return {
            "status": 400,
            "message": str(e)
        }

    return {
        "status": 200,
        "message": "Message was successfully added to queue. The process is running on background.",
        "messageId": response['MessageId']
    }


@app.on_sqs_message(queue=queue_name, batch_size=1)
def handle_sqs_message(event):
    for record in event:
        app.log.info("Received message with contents: %s", record.body)
        reporter = ReportSplitter(
            app=app, s3=s3, bucket_name=bucket_name,
            input_template_key=input_template_key, sendgrid_key=sendgrid_key, bugsnag_key=bugsnag_key)

        input_dict = json.loads(record.body)
        reporter.process(input_dict)


# '''
# tester, local
@app.route('/test-new')
def index_new():
    app.log.info("Handler triggered.....")

    input_dict_csv = {
        "report_name": "ali_test",
        "org_id": "test_org",
        "course_name": "test_course_name",
        "course_id": "test_course_id",
        "filter_message": "test_filter_message",
        "generated_timestamp": "2021-1-22",
        "email_addresses": ["upworka0@gmail.com"],
        "email_sender": "upworka01@gmail.com",
        "group_by": ["ref_State", "ref_City"],
        "CSV_filepath": "sample_03.csv",
        "email_subject": "6 Apr testing"
    }
    # return reporter.process(input_dict_csv)

    input_dict_zip = {
        "report_name": "Maine test-20211003",
        "org_id": 442,
        "filter_message": "",
        "generated_timestamp": "03-Oct-2021",
        "email_addresses": [
            "johannes@getinclusive.com"
        ],
        "email_sender": "Get Inclusive <support@getinclusive.com>",
        "email_subject": "STAGING - Progress Report - as of 03-Oct-2021 - johannes@getinclusive.com",
        "ZIP_filepath": "2021-10-12-11-02-30.zip",
        "courses": {
            "course_2014.csv": {
                "course_name": "Harassment Prevention/Title IX Education Abridged (CA)",
                "course_id": 2014
            }
        }
    }

    # queue = sqs.get_queue_by_name(QueueName=queue_name)
    # response = queue.send_message(MessageBody=json.dumps(input_dict_zip))
    # app.log.info("Message was created with %s" % input_dict_zip)

    reporter = ReportSplitter(
        app=app, s3=s3, bucket_name=bucket_name,
        input_template_key=input_template_key, sendgrid_key=sendgrid_key, bugsnag_key=bugsnag_key)

    reporter.process(input_dict_zip)

    return {
        "status": 200,
        "message": "Message was successfully added to queue. The process is running on background.",
        # "messageId": response['MessageId']
    }
